from __future__ import annotations

import json
import mimetypes
import sys
from io import BytesIO
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from threading import Lock
from urllib.parse import unquote, urlparse

ROOT = Path(__file__).resolve().parent.parent
VISUALS_DIR = ROOT / "07. Visutals"
INDEX_PATH = VISUALS_DIR / "index.html"
VIEWS_PATH = ROOT / "05. Cost modeling" / ".cost-model-views.json"
HOST = "127.0.0.1"
PORT = 8000

sys.path.insert(0, str(ROOT / "05. Cost modeling"))
from generate_dashboard import dashboard_table_data, generate_dashboard, product_records, read_position_statuses  # noqa: E402

DEFAULT_VIEWS = {
    "current": {"visible": list(range(10)), "filters": [], "groups": [], "sorts": []},
    "position-size": {"visible": list(range(10)), "filters": [], "groups": [], "sorts": []},
}
views_lock = Lock()
FIXED_VIEW_NAMES = {"90. Definitions", "91. Sources", "92. Data download"}
LEGACY_FIXED_VIEW_NAMES = {"Definitions", "Sources"}


def normalize_saved_views(value: dict) -> dict:
    normalized = {
        name: state
        for name, state in value.items()
        if name not in FIXED_VIEW_NAMES and name not in LEGACY_FIXED_VIEW_NAMES
    }
    return normalized or DEFAULT_VIEWS.copy()


def workbook_bytes() -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    tables = dashboard_table_data(product_records(), read_position_statuses())
    workbook = Workbook()
    workbook.remove(workbook.active)
    sheets = (
        ("Data", tables["data_headers"], tables["data_rows"]),
        ("Sources", tables["source_headers"], tables["source_rows"]),
        ("Definitions", tables["definition_headers"], tables["definition_rows"]),
    )
    for title, headers, rows in sheets:
        worksheet = workbook.create_sheet(title)
        worksheet.append(list(headers))
        for row in rows:
            worksheet.append(list(row))
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        worksheet.sheet_view.showGridLines = False
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(fill_type="solid", fgColor="176B52")
        for index, column in enumerate(worksheet.iter_cols(), start=1):
            longest = max((len(str(cell.value or "")) for cell in column), default=0)
            worksheet.column_dimensions[get_column_letter(index)].width = min(max(longest + 2, 12), 60)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def read_views() -> dict:
    if not VIEWS_PATH.exists():
        return DEFAULT_VIEWS.copy()
    try:
        value = json.loads(VIEWS_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return DEFAULT_VIEWS.copy()
    return normalize_saved_views(value) if isinstance(value, dict) and value else DEFAULT_VIEWS.copy()


def write_views(value: dict) -> None:
    with views_lock:
        temporary_path = VIEWS_PATH.with_suffix(".tmp")
        temporary_path.write_text(json.dumps(value, indent=2) + "\n", encoding="utf-8")
        temporary_path.replace(VIEWS_PATH)


class CostModelHandler(BaseHTTPRequestHandler):
    def send_json(self, status: HTTPStatus, value: object) -> None:
        payload = json.dumps(value).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(payload)))
        self.end_headers()
        self.wfile.write(payload)

    def do_GET(self) -> None:
        path = urlparse(self.path).path
        if path == "/api/views":
            self.send_json(HTTPStatus.OK, read_views())
            return
        if path == "/api/downloads/data-workbook.xlsx":
            try:
                payload = workbook_bytes()
            except Exception as error:
                self.send_json(HTTPStatus.INTERNAL_SERVER_ERROR, {"error": str(error)})
                return
            self.send_response(HTTPStatus.OK)
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition", 'attachment; filename="structured-products-data.xlsx"')
            self.send_header("Content-Length", str(len(payload)))
            self.end_headers()
            self.wfile.write(payload)
            return
        if path == "/":
            path = "/index.html"
        requested = (VISUALS_DIR / unquote(path.lstrip("/"))).resolve()
        if requested != INDEX_PATH.resolve() or not requested.is_file():
            self.send_error(HTTPStatus.NOT_FOUND)
            return
        generate_dashboard()
        payload = requested.read_bytes()
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", mimetypes.guess_type(str(requested))[0] or "text/plain")
        self.send_header("Content-Length", str(len(payload)))
        self.end_headers()
        self.wfile.write(payload)

    def do_PUT(self) -> None:
        if urlparse(self.path).path != "/api/views":
            self.send_error(HTTPStatus.NOT_FOUND)
            return
        try:
            length = int(self.headers.get("Content-Length", "0"))
            value = json.loads(self.rfile.read(length))
        except (ValueError, json.JSONDecodeError):
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "Request body must be valid JSON."})
            return
        if not isinstance(value, dict) or not value:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "At least one view is required."})
            return
        value = normalize_saved_views(value)
        try:
            write_views(value)
        except OSError:
            self.send_json(HTTPStatus.INTERNAL_SERVER_ERROR, {"error": "View changes could not be saved."})
            return
        self.send_json(HTTPStatus.OK, value)

    def log_message(self, format: str, *args: object) -> None:
        print(f"{self.address_string()} - {format % args}")


def main() -> None:
    server = ThreadingHTTPServer((HOST, PORT), CostModelHandler)
    print(f"Cost model dashboard: http://{HOST}:{PORT}/")
    print(f"View data: {VIEWS_PATH.relative_to(ROOT)}")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nStopping cost model server")
    finally:
        server.server_close()


if __name__ == "__main__":
    main()
