from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
WORKBOOK_PATH = ROOT / "Trust ISIN information from Bloomberg.xlsx"
OCR_DIR = ROOT / "03. BBG OCR"
OUTPUT_PATH = ROOT / "04. Product Review" / "ISSUE SIZE RECOVERY.md"

ISIN_PATTERN = re.compile(r"^(?:XS|CH)\d{10}$")
LABEL_PATTERN = re.compile(
    r"(?i)\b(?:amt\s*issued\s*/\s*outstanding|amount\s+issued\s*/\s*outstanding|"
    r"amount\s+issued|issue\s+size|aggregate\s+(?:nominal|principal|amount)|"
    r"nominal\s+amount|principal\s+amount|outstanding\s+amount)\b"
)
AMOUNT_PATTERN = re.compile(
    r"(?i)\b(USD|EUR|CHF|GBP|AUD|CAD|JPY)\s*\n?\s*([\d,]+(?:\.\d+)?)\s*\(M\)"
)


def portfolio_isins() -> list[str]:
    workbook = load_workbook(WORKBOOK_PATH, data_only=True, read_only=True)
    worksheet = workbook["ISINs summary"]
    isins = []
    for row in worksheet.iter_rows(values_only=True):
        for value in row:
            if isinstance(value, str) and ISIN_PATTERN.fullmatch(value.strip()):
                if value.strip() not in isins:
                    isins.append(value.strip())
    return isins


def recover_from_ocr(isin: str) -> list[tuple[Path, str, str]]:
    findings = []
    for path in sorted((OCR_DIR / isin).glob("*.txt")):
        lines = path.read_text(encoding="utf-8", errors="replace").splitlines()
        for index, line in enumerate(lines):
            match = LABEL_PATTERN.search(line)
            if not match:
                continue
            context_lines = [line.strip()]
            for candidate in lines[index + 1 : index + 12]:
                candidate = candidate.strip()
                if candidate:
                    context_lines.append(candidate)
            context = " ".join(context_lines)
            amount = AMOUNT_PATTERN.search("\n".join(lines[index + 1 : index + 12]))
            if amount:
                value = f"{amount.group(1).upper()} {amount.group(2)} million"
                context = f"{context}; recovered numeric candidate: {value}"
            else:
                context = f"{context}; no numeric value parsed"
            findings.append((path, match.group(0), context))
    return findings


def main() -> None:
    rows = []
    confirmed = 0
    candidates = 0
    unavailable = 0
    for isin in portfolio_isins():
        findings = recover_from_ocr(isin)
        if findings:
            candidates += 1
            evidence = "<br>".join(
                f"`{path.relative_to(ROOT).as_posix()}`: {context}"
                for path, _, context in findings
            )
            status = "Candidate: visual confirmation required"
        else:
            unavailable += 1
            evidence = "No issue-size label found in available OCR"
            status = "Unavailable: document recovery required"
        rows.append((isin, len(findings), evidence, status))

    lines = [
        "# Issue Size Recovery",
        "",
        "> Automated recovery pass over all portfolio ISINs and available Bloomberg OCR.",
        "> `Amt Issued/outstanding` is treated as a candidate issue/outstanding-size field, not confirmed evidence until checked against the source image.",
        "",
        "## Results",
        "",
        "| ISIN | OCR labels found | Recovered field/context | Status |",
        "| --- | ---: | --- | --- |",
    ]
    lines.extend(f"| `{isin}` | {count} | {evidence} | {status} |" for isin, count, evidence, status in rows)
    lines.extend([
        "",
        "## Totals",
        "",
        f"- Portfolio instruments scanned: {len(rows)}",
        f"- Instruments with issue-size label candidates: {candidates}",
        f"- Instruments with no available issue-size label: {unavailable}",
        f"- Confirmed original issue sizes: {confirmed}",
        "",
        "## Interpretation",
        "",
        "- The recovery pass does not convert the existing Trust position-size field into issuance size.",
        "- OCR candidates must be visually checked against the linked Bloomberg image and recorded with the displayed currency and amount.",
        "- Instruments without a candidate require final terms, pricing supplements, listing records, or issuer/custodian recovery.",
    ])
    OUTPUT_PATH.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"Scanned {len(rows)} instruments")
    print(f"Issue-size label candidates: {candidates}")
    print(f"No issue-size label found: {unavailable}")
    print(f"Wrote {OUTPUT_PATH}")


if __name__ == "__main__":
    main()