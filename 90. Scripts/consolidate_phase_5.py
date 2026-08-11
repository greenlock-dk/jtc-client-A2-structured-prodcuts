from __future__ import annotations

import os
import re
from collections import Counter
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from generate_phase_1_2 import format_value
from product_frontmatter import read_product, write_product
from review_phase_4b import dossier_for, find_candidate

ROOT = Path(__file__).resolve().parent.parent
PRODUCT_DIR = ROOT / "01. Structured Products"
EVIDENCE_DIR = ROOT / "02. BBG images"
OCR_DIR = ROOT / "03. BBG OCR"
REVIEW_DIR = ROOT / "04. Product Review"
WORKBOOK_PATH = ROOT / "Trust ISIN information from Bloomberg.xlsx"
INVENTORY_PATH = REVIEW_DIR / "PHASE 5 CONSOLIDATION.md"

EVIDENCE_RESOLUTIONS = {
    ("XS0164480286", "Maturity / call date"): (
        "Maturity: 21-Mar-2018; Call effective: 25-Mar-2008",
        "Bloomberg image 01 shows maturity 03/21/2018 and call effective date 03/25/2008; "
        "the workbook value 25-Mar-2018 matches neither labeled date.",
    ),
    ("XS0241444883", "Coupon / yield"): (
        "Initial 7.25%; ratchet increments of 0.05%, 0.15%, and 0.25% per the dated coupon schedule",
        "Bloomberg image 03 shows an initial fixed 7.25% coupon followed by dated ratchet formulas; "
        "the workbook's 7.25-7.5% range is an incomplete summary.",
    ),
    ("XS0249805960", "Maturity / call date"): (
        "Maturity: 11-May-2021",
        "Bloomberg image 01 shows maturity 05/11/2021; the workbook records 11-Mar-2021.",
    ),
}

FIELDS = (
    "Source / exhibit",
    "Trust",
    "Country of instrument",
    "Custodian / bank held",
    "ISIN",
    "Issuer",
    "Product name",
    "Product type / structure",
    "Currency",
    "Issue date",
    "Maturity / call date",
    "Tenor (years)",
    "Underlying(s)",
    "Coupon / yield",
    "Barrier / protection level",
    "Observation / payment frequency",
    "Annualised rate",
    "Other comments",
    "Redemption terms",
    "Downside",
    "Risk notes",
    "Position size (USD)",
    "Denomination (USD)",
)

OCR_FIELD_MAP = {
    "ISIN": "ISIN",
    "Issuer": "Issuer",
    "Currency": "Currency",
    "Maturity / call date": "Maturity",
    "Coupon / yield": "Coupon / yield",
    "Product type / structure": "Product type / structure",
    "Underlying(s)": "Underlying(s)",
}


def clean(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("|", "\\|").replace("\n", "<br>").strip()


def normalize(value: str) -> str:
    return re.sub(r"[^a-z0-9]", "", value.lower())


def parsed_dates(value: str) -> set[str]:
    dates = set()
    for pattern, date_format in (
        (r"\b\d{2}-[A-Za-z]{3}-\d{4}\b", "%d-%b-%Y"),
        (r"\b\d{2}/\d{2}/\d{4}\b", "%m/%d/%Y"),
    ):
        for candidate in re.findall(pattern, value):
            try:
                dates.add(datetime.strptime(candidate, date_format).date().isoformat())
            except ValueError:
                continue
    return dates


def percentage_values(value: str) -> set[float]:
    return {
        round(float(candidate), 6)
        for candidate in re.findall(r"(?<![\d.])(\d+(?:\.\d+)?)\s*%", value)
        if float(candidate) <= 100
    }


def issuer_tokens(value: str) -> set[str]:
    normalized = value.lower()
    for pattern, replacement in (
        (r"\bbanq\b", "banque"),
        (r"\bintl\b", "international"),
        (r"\baust\b", "australia"),
    ):
        normalized = re.sub(pattern, replacement, normalized)
    return set(re.findall(r"[a-z]+", normalized)) - {"plc", "sa"}


def relative_link(path: Path, base: Path) -> str:
    return Path(os.path.relpath(path, base)).as_posix().replace(" ", "%20")


def worksheet_records(workbook) -> dict[str, dict[str, str]]:
    worksheet = workbook["ISINs summary"]
    records = {}
    recent_columns = {
        "Source / exhibit": 2,
        "Trust": 3,
        "ISIN": 4,
        "Product name": 5,
        "Issue date": 6,
        "Maturity / call date": 7,
        "Tenor (years)": 8,
        "Product type / structure": 9,
        "Coupon / yield": 10,
        "Observation / payment frequency": 11,
        "Annualised rate": 12,
        "Barrier / protection level": 13,
        "Underlying(s)": 14,
        "Other comments": 15,
        "Redemption terms": 16,
        "Risk notes": 17,
        "Position size (USD)": 18,
        "Denomination (USD)": 19,
    }
    historical_columns = {
        "Source / exhibit": 2,
        "Trust": 3,
        "ISIN": 4,
        "Country of instrument": 5,
        "Custodian / bank held": 6,
        "Issuer": 7,
        "Issue date": 8,
        "Maturity / call date": 9,
        "Tenor (years)": 10,
        "Product type / structure": 11,
        "Coupon / yield": 12,
        "Observation / payment frequency": 13,
        "Annualised rate": 14,
        "Barrier / protection level": 15,
        "Underlying(s)": 16,
        "Other comments": 17,
        "Redemption terms": 18,
        "Risk notes": 19,
        "Position size (USD)": 20,
        "Denomination (USD)": 21,
    }
    for row_number in range(7, 16):
        isin = format_value(worksheet.cell(row_number, 4))
        if isin.startswith(("XS", "CH")):
            records[isin] = {
                field: format_value(worksheet.cell(row_number, column))
                for field, column in recent_columns.items()
            }
            records[isin]["Downside"] = records[isin].pop("Redemption terms", "")
    for row_number in range(20, worksheet.max_row + 1):
        isin = format_value(worksheet.cell(row_number, 4))
        if isin.startswith(("XS", "CH")):
            records[isin] = {
                field: format_value(worksheet.cell(row_number, column))
                for field, column in historical_columns.items()
            }
            records[isin]["Product name"] = records[isin]["Product type / structure"]
            records[isin]["Downside"] = records[isin].pop("Redemption terms", "")
    return records


def ocr_fields(isin: str) -> dict[str, str]:
    texts = sorted((OCR_DIR / isin).glob("*.txt"))
    lines = []
    for path in texts:
        lines.extend(line.strip() for line in path.read_text(encoding="utf-8").splitlines() if line.strip())
    combined = "\n".join(lines)
    issuer = find_candidate(lines, "Issuer") or find_candidate(lines, "Name")
    currency = find_candidate(lines, "Currency")
    maturity = find_candidate(lines, "Maturity", r"Maturity\s+(\d{2}/\d{2}/\d{4})")
    call_effective = find_candidate(lines, "Call Effective Date", r"Call Effective Date\s+(\d{2}/\d{2}/\d{4})")
    coupons = re.findall(r"\b\d+(?:\.\d+)?%", combined)
    structures = []
    for term in (
        "Range Accrual", "Dual Range Accrual", "Dual Index", "Ratchet",
        "Reverse Convertible", "Convertible Bond", "Perpetual", "EURO MTN", "EMTN",
    ):
        if term.lower() in combined.lower() and term not in structures:
            structures.append(term)
    return {
        "ISIN": isin if re.search(re.escape(isin), combined, re.IGNORECASE) else "",
        "Issuer": issuer,
        "Currency": currency,
        "Maturity / call date": "; ".join(
            value for value in (
                f"Maturity: {maturity}" if maturity else "",
                f"Call effective: {call_effective}" if call_effective else "",
            ) if value
        ),
        "Coupon / yield": ", ".join(dict.fromkeys(coupons)),
        "Product type / structure": ", ".join(structures),
        "Underlying(s)": "",
    }


def source_files(isin: str) -> tuple[list[Path], list[Path]]:
    images = sorted((EVIDENCE_DIR / isin).glob("*.png"))
    ocr = sorted((OCR_DIR / isin).glob("*.txt"))
    return images, ocr


def ocr_identifier_issues(isin: str, ocr_paths: list[Path]) -> list[str]:
    issues = []
    identifiers = set()
    for path in ocr_paths:
        text = path.read_text(encoding="utf-8")
        identifiers.update(re.findall(r"\b(?:XS|CH)[A-Z0-9]{10}\b", text, re.IGNORECASE))
    for identifier in sorted(identifiers):
        if identifier.upper() != isin.upper():
            issues.append(f"HIGH: OCR contains identifier {identifier}, not {isin}")
    return issues


def comparison(field: str, workbook_value: str, ocr_value: str) -> str:
    if not workbook_value or not ocr_value:
        return "unavailable"
    if normalize(workbook_value) == normalize(ocr_value):
        return "corroborated"
    if field == "Maturity / call date":
        workbook_dates = parsed_dates(workbook_value)
        if workbook_dates and workbook_dates & parsed_dates(ocr_value):
            return "corroborated"
    if field == "Issuer":
        workbook_tokens = issuer_tokens(workbook_value)
        ocr_tokens = issuer_tokens(ocr_value)
        if workbook_tokens and (workbook_tokens <= ocr_tokens or ocr_tokens <= workbook_tokens):
            return "corroborated"
    if field == "Coupon / yield":
        workbook_rates = percentage_values(workbook_value)
        if workbook_rates and workbook_rates <= percentage_values(ocr_value):
            return "corroborated"
    if field == "Product type / structure":
        return "complementary"
    return "conflict"


def consolidated_rows(isin: str, workbook_record: dict[str, str], candidates: dict[str, str]):
    rows = []
    discrepancies = []
    resolutions = []
    for field in FIELDS:
        workbook_value = clean(workbook_record.get(field, ""))
        ocr_value = clean(candidates.get(field, ""))
        result = comparison(field, workbook_value, ocr_value)
        resolution = EVIDENCE_RESOLUTIONS.get((isin, field))
        if result == "conflict" and resolution:
            consolidated, rationale = resolution
            status = "Resolved by visual Bloomberg review"
            resolutions.append(f"{field}: {rationale}")
        elif result == "conflict":
            severity = "HIGH" if field in {"ISIN", "Maturity / call date"} else "MEDIUM"
            discrepancies.append(
                f"{severity}: workbook and OCR differ for {field} "
                f"(workbook: {workbook_value}; OCR: {ocr_value})"
            )
        if resolution and result == "conflict":
            pass
        elif workbook_value:
            consolidated = workbook_value
            if result == "corroborated":
                status = "Corroborated by OCR"
            elif result == "complementary":
                status = "Complementary source descriptions"
            elif result == "conflict":
                status = "Workbook reference; OCR conflict"
            else:
                status = "Workbook reference"
        elif ocr_value:
            consolidated = ocr_value
            status = "OCR candidate; visual confirmation required"
        else:
            consolidated = ""
            status = "Not available in current sources"
        rows.append((field, workbook_value, ocr_value, consolidated, status))
    return rows, discrepancies, resolutions


def write_dossier(isin: str, workbook_record: dict[str, str]) -> tuple[int, int, int, int]:
    dossier = dossier_for(isin)
    canonical_data, _ = read_product(dossier) if dossier.exists() else ({}, "")
    images, ocr_paths = source_files(isin)
    source_count = len(images)
    ocr_count = len(ocr_paths)
    candidates = ocr_fields(isin)
    rows, discrepancies, resolutions = consolidated_rows(isin, workbook_record, candidates)
    discrepancies.extend(ocr_identifier_issues(isin, ocr_paths))
    if source_count != ocr_count:
        discrepancies.append(
            f"HIGH: evidence image count ({source_count}) differs from OCR file count ({ocr_count})"
        )
    image_links = [
        f"[{path.name}]({relative_link(path, dossier.parent)})" for path in images
    ]
    ocr_links = [
        f"[{path.name}]({relative_link(path, dossier.parent)})" for path in ocr_paths
    ]
    if source_count == ocr_count and source_count > 0:
        source_status = "Complete"
    elif source_count == ocr_count == 0:
        source_status = "Source gap: no extracted evidence"
    else:
        source_status = "Inconsistent: image/OCR counts differ"
    lines = [
        f"# {dossier.stem}",
        "",
        f"- ISIN: `{isin}`",
        "- Source workbook: `Trust ISIN information from Bloomberg.xlsx`, worksheet `ISINs summary`",
        "- Consolidation phase: Phase 5",
        f"- Evidence images: {source_count}; OCR files: {ocr_count}; source status: {source_status}",
        "",
        "## Consolidated Product Record",
        "",
        "Workbook values are retained as the reference layer. OCR values are supporting candidates and do not replace workbook values without visual confirmation.",
        "",
        "| Field | Workbook reference | OCR candidate | Consolidated working value | Status |",
        "| --- | --- | --- | --- | --- |",
    ]
    lines.extend(
        f"| {field} | {workbook_value} | {ocr_value} | {consolidated} | {status} |"
        for field, workbook_value, ocr_value, consolidated, status in rows
    )
    lines.extend([
        "",
        "## Consistency Review",
        "",
        "Disagreements are preserved for manual source-image review; no unverified OCR correction is applied.",
        "",
    ])
    if discrepancies:
        lines.extend(f"- {item}" for item in discrepancies)
    else:
        lines.append("- No unresolved workbook/OCR field disagreements detected.")
    if resolutions:
        lines.extend(["", "Resolved through visual review:", ""])
        lines.extend(f"- {item}" for item in resolutions)
    lines.extend([
        "",
        "## Evidence Sources",
        "",
        f"- Images: {', '.join(image_links) if image_links else 'None available'}",
        f"- Raw OCR: {', '.join(ocr_links) if ocr_links else 'None available'}",
        "",
        "## Review Rules",
        "",
        "- `Workbook reference` is source data from the structured workbook and is not a confirmation of contractual terms.",
        "- `OCR candidate` is a machine transcription and requires visual comparison with the linked image.",
        "- A disagreement is not resolved until the source image or an original term sheet is checked.",
    ])
    # Reviewed frontmatter is canonical and must survive an evidence-report rebuild.
    if canonical_data:
        write_product(dossier, canonical_data, "\n".join(lines) + "\n")
    else:
        dossier.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return source_count, ocr_count, len(discrepancies), len(resolutions)


def main() -> None:
    REVIEW_DIR.mkdir(exist_ok=True)
    workbook = load_workbook(WORKBOOK_PATH, data_only=True)
    records = worksheet_records(workbook)
    inventory = [
        "# Phase 5 Consolidation",
        "",
        "> Consolidated workbook, OCR, and extracted-image references for each instrument.",
        "> Workbook values remain the reference layer; OCR candidates remain explicitly unconfirmed.",
        "",
        "| Record | ISIN | Images | OCR files | Unresolved | Resolved | Status |",
        "| ---: | --- | ---: | ---: | ---: | ---: | --- |",
    ]
    totals = Counter()
    for record_number, isin in enumerate(sorted(records), 1):
        image_count, ocr_count, discrepancy_count, resolution_count = write_dossier(isin, records[isin])
        totals.update({
            "instruments": 1,
            "images": image_count,
            "ocr": ocr_count,
            "discrepancies": discrepancy_count,
            "resolutions": resolution_count,
        })
        if discrepancy_count:
            status = "Review required"
        elif resolution_count:
            status = "Resolved by visual review"
        elif image_count == 0 and ocr_count == 0:
            status = "Source gap; no cross-source comparison"
        else:
            status = "No detected inconsistency"
        inventory.append(
            f"| {record_number} | `{isin}` | {image_count} | {ocr_count} | "
            f"{discrepancy_count} | {resolution_count} | {status} |"
        )
    inventory.extend([
        "",
        "## Totals",
        "",
        f"- Instruments consolidated: {totals['instruments']}",
        f"- Evidence images linked: {totals['images']}",
        f"- OCR files linked: {totals['ocr']}",
        f"- Unresolved workbook/OCR disagreements: {totals['discrepancies']}",
        f"- Disagreements resolved by visual review: {totals['resolutions']}",
        "",
        "## Interpretation",
        "",
        "Phase 5 consolidates available information; it does not certify product terms. Manual visual review remains necessary wherever a conflict or OCR-only candidate is present.",
    ])
    INVENTORY_PATH.write_text("\n".join(inventory) + "\n", encoding="utf-8")
    print(f"Consolidated {totals['instruments']} instruments")
    print(f"Linked {totals['images']} images and {totals['ocr']} OCR files")
    print(f"Recorded {totals['discrepancies']} unresolved workbook/OCR disagreements")
    print(f"Recorded {totals['resolutions']} visual-review resolutions")
    print(f"Created {INVENTORY_PATH}")


if __name__ == "__main__":
    main()