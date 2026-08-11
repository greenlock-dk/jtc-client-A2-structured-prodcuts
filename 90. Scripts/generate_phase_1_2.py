from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
import re
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell

ROOT = Path(__file__).resolve().parent.parent
WORKBOOK_PATH = ROOT / "Trust ISIN information from Bloomberg.xlsx"
OUTPUT_DIR = ROOT / "01. Structured Products"
CANONICAL_DIR = ROOT / "05. Canonical Data"
DETAILED_PATH = CANONICAL_DIR / "ISIN_detailed.md"
SUMMARY_PATH = CANONICAL_DIR / "ISIN_summary.md"
NON_ISIN_SHEETS = {"Jogger - Tudella 30 June 2009", "ISINs summary", "Bloomberg data >>"}
PRODUCT_SUFFIXES = {
    "XS0765564827": "Aquarius Secured Notes",
    "XS1028242706": "Morgan Stanley EMTN",
    "XS1243914071": "Nomura Euro Dollar",
    "CH0252328973": "Credit Suisse MTN",
    "XS0297701319": "Callable Range Note",
    "XS0300388351": "Dual Index Note",
    "XS0164480286": "Libor Callable Note",
    "XS0165220400": "Libor Range Note",
    "XS0169318291": "Libor Callable Note",
    "XS0170303290": "Libor Callable Note",
    "XS0171914038": "Libor Callable Note",
    "XS0172077769": "Libor Callable Note",
    "XS0241444883": "Libor Range Note",
    "XS0249805960": "Libor Callable Note",
    "XS0277502067": "Libor Callable Note",
    "XS0278550750": "Libor Callable Note",
    "XS0284203071": "CMS Spread Note",
    "XS0294314694": "Libor Callable Note",
    "XS0293931688": "CMS Spread Note",
    "XS0293919121": "Libor Callable Note",
    "XS0297467705": "Libor Callable Note",
    "XS0304286924": "Libor Callable Note",
    "XS0314283432": "Callable Note",
    "XS0315745447": "Fixed Rate Note",
    "CH1484588913": "Leonteq Express Certificate",
    "XS3234638248": "BBVA Phoenix Memory",
    "XS0168875792": "Libor Callable Note",
    "XS0298465822": "Unspecified Instrument",
    "XS0318585791": "Kick In Note",
}


def is_isin_sheet(sheet_name: str) -> bool:
    normalized_name = sheet_name.strip()
    return normalized_name not in NON_ISIN_SHEETS and normalized_name.startswith(("XS", "CH"))


def clean_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("|", "\\|").replace("\r\n", "\n").replace("\r", "\n").replace("\n", "<br>").strip()


def format_value(cell: Cell) -> str:
    value = cell.value
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.strftime("%d-%b-%Y")
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        fmt = cell.number_format or ""
        if "%" in fmt:
            decimals = len(fmt.split(".", 1)[1].replace("%", "")) if "." in fmt else 0
            return f"{value * 100:.{decimals}f}%"
        if "#" in fmt and "." not in fmt:
            return f"{value:,.0f}"
        return f"{value:g}"
    return clean_text(value)


def row_is_empty(cells: Iterable[Cell]) -> bool:
    return not any(cell.value is not None for cell in cells)


DETAILED_HEADERS = [
    "Record #", "Source section", "Source / Exhibit", "Trust", "ISIN", "Name of product",
    "Country of instrument", "Bank held", "Issuer", "Issue date", "Maturity",
    "Tenor (years)", "Structure", "Coupon rate", "frequency", "Annualised rate",
    "Barrier", "underlying", "Other comments", "Downside", "Risk", "Size (USD)",
    "Denomination (USD)", "Summary link",
]


def markdown_table(headers: list[str], rows: list[list[str]]) -> str:
    lines = ["| " + " | ".join(headers) + " |", "| " + " | ".join("---" for _ in headers) + " |"]
    lines.extend("| " + " | ".join(row) + " |" for row in rows)
    return "\n".join(lines)


def detailed_rows(ws) -> list[list[str]]:
    rows = []
    record_number = 1
    for row_number in range(7, 16):
        cells = [ws.cell(row_number, col) for col in range(1, 20)]
        if row_is_empty(cells):
            continue
        isin = format_value(ws.cell(row_number, 4))
        rows.append([
            f'<a id="record-{record_number}"></a>{record_number}', "Recent structured product", format_value(ws.cell(row_number, 2)), format_value(ws.cell(row_number, 3)),
            isin, format_value(ws.cell(row_number, 5)), "", "", "", format_value(ws.cell(row_number, 6)),
            format_value(ws.cell(row_number, 7)), format_value(ws.cell(row_number, 8)), format_value(ws.cell(row_number, 9)),
            format_value(ws.cell(row_number, 10)), format_value(ws.cell(row_number, 11)), format_value(ws.cell(row_number, 12)),
            format_value(ws.cell(row_number, 13)), format_value(ws.cell(row_number, 14)), format_value(ws.cell(row_number, 15)),
            format_value(ws.cell(row_number, 16)), format_value(ws.cell(row_number, 17)), format_value(ws.cell(row_number, 18)),
            format_value(ws.cell(row_number, 19)), f"[Summary](ISIN_summary.md#record-{record_number})",
        ])
        record_number += 1
    for row_number in range(20, ws.max_row + 1):
        cells = [ws.cell(row_number, col) for col in range(1, 22)]
        if row_is_empty(cells):
            continue
        isin = format_value(ws.cell(row_number, 4))
        structure = format_value(ws.cell(row_number, 11))
        rows.append([
            f'<a id="record-{record_number}"></a>{record_number}', "Historical portfolio instrument", format_value(ws.cell(row_number, 2)), format_value(ws.cell(row_number, 3)),
            isin, structure, format_value(ws.cell(row_number, 5)), format_value(ws.cell(row_number, 6)),
            format_value(ws.cell(row_number, 7)), format_value(ws.cell(row_number, 8)), format_value(ws.cell(row_number, 9)),
            format_value(ws.cell(row_number, 10)), structure, format_value(ws.cell(row_number, 12)), format_value(ws.cell(row_number, 13)),
            format_value(ws.cell(row_number, 14)), format_value(ws.cell(row_number, 15)), format_value(ws.cell(row_number, 16)),
            format_value(ws.cell(row_number, 17)), format_value(ws.cell(row_number, 18)), format_value(ws.cell(row_number, 19)),
            format_value(ws.cell(row_number, 20)), format_value(ws.cell(row_number, 21)), f"[Summary](ISIN_summary.md#record-{record_number})",
        ])
        record_number += 1
    return rows


def write_summary(workbook) -> None:
    ws = workbook["ISINs summary"]
    tab_isins = {name.strip() for name in workbook.sheetnames if is_isin_sheet(name)}
    summary_isins = {
        str(ws.cell(row, 4).value).strip()
        for row in range(1, ws.max_row + 1)
        if ws.cell(row, 4).value and is_isin_sheet(str(ws.cell(row, 4).value).strip())
    }
    missing_tabs = sorted(summary_isins - tab_isins)
    rows = detailed_rows(ws)
    detailed_lines = [
        "# ISIN Detailed",
        "",
        "> Source: `Trust ISIN information from Bloomberg.xlsx`, worksheet `ISINs summary`.",
        "> Generated from the workbook using `openpyxl`; dates and percentages follow the workbook display formats.",
        "",
        "## Project Context",
        "",
        "> I realize it is just ISINs with Bloomberg screen shot of term sheets. But it is the initial major gap of information we have in our process.",
        ">",
        "> I am fully aware your scope is much larger than just autocallable products and retrocessions. But this is an area that we could really use a Swiss focused expert.",
        ">",
        "> Note: the top section of the first tab is more recent in the last decade. The bottom section of the first tab is about 19 years old. Not recoverable from banks but still relevant for our fiduciary investigation.",
        ">",
        "> Do you think you could find some of the term sheets.",
        "",
        "## Flattened Product Records",
        "",
        "Recent and historical records are combined below. The `Source section` column preserves the original workbook section, and `Summary link` cross-references each record to `ISIN_summary.md`.",
        "",
        markdown_table(DETAILED_HEADERS, rows),
        "",
        "## Worksheet Coverage",
        "",
        f"- ISIN worksheets found: {len(tab_isins)}",
        f"- ISINs listed in the summary: {len(summary_isins)}",
        f"- Summary ISINs without a dedicated worksheet: {len(missing_tabs)}",
    ]
    if missing_tabs:
        detailed_lines.extend(["", "The following summary ISINs do not have a dedicated worksheet in the source workbook:", ""])
        detailed_lines.extend(f"- `{isin}`" for isin in missing_tabs)
    detailed_lines.extend([
        "",
        "## Interpretation Notes",
        "",
        "- Blank cells are preserved as blank values; they are not interpreted as zero or as unavailable evidence.",
        "- Formula cells are rendered using their cached Excel results where available.",
        "- Historical source rows do not provide a dedicated product-name field; their recorded structure is used as `Name of product`, while the original country remains in `Country of instrument`.",
        "- Image extraction and OCR are intentionally deferred to a later phase.",
    ])
    DETAILED_PATH.write_text("\n".join(detailed_lines) + "\n", encoding="utf-8")

    summary_headers = ["Record #", "Name of product", "ISIN", "Maturity", "Size (USD)", "Detailed record"]
    summary_rows = []
    for row in rows:
        record_number = re.search(r'record-(\d+)', row[0]).group(1)
        summary_rows.append([row[0], row[5], row[4], row[10], row[21], f"[Detailed](ISIN_detailed.md#record-{record_number})"])
    short_lines = [
        "# ISIN Summary",
        "",
        "> Source: `Trust ISIN information from Bloomberg.xlsx`, worksheet `ISINs summary`.",
        "> This short table is generated from the same normalized records as [ISIN_detailed.md](ISIN_detailed.md), ensuring both tables remain synchronized.",
        "",
        markdown_table(summary_headers, summary_rows),
        "",
        f"- Records: {len(summary_rows)}",
        f"- ISIN worksheets found: {len(tab_isins)}",
        f"- Summary ISINs without a dedicated worksheet: {len(missing_tabs)}",
    ]
    SUMMARY_PATH.write_text("\n".join(short_lines) + "\n", encoding="utf-8")


def write_isin_placeholder(workbook, isin: str) -> bool:
    has_worksheet = isin in workbook.sheetnames
    ws = workbook[isin] if has_worksheet else None
    image_count = len(ws._images) if ws else 0
    product_suffix = PRODUCT_SUFFIXES[isin]
    path = OUTPUT_DIR / f"{isin} - {product_suffix}.md"
    if path.exists():
        return False
    source_worksheet = f"`{isin}`" if has_worksheet else "No dedicated worksheet in source workbook"
    processing_status = (
        "Placeholder created in Phases 1-2; image extraction and OCR pending."
        if has_worksheet
        else "Summary-only record: no dedicated worksheet, embedded images, or OCR evidence is available in the source workbook."
    )
    content = f"""# {isin} - {product_suffix}

- Source worksheet: {source_worksheet}
- Source workbook: `Trust ISIN information from Bloomberg.xlsx`
- Embedded images currently identified on worksheet: {image_count}
- Processing status: {processing_status}

## Product Information

| Field | Value | Status |
| --- | --- | --- |
| ISIN | `{isin}` | From worksheet name |
| Issuer |  | Pending image and source review |
| Product name |  | Pending image and source review |
| Product type / structure |  | Pending image and source review |
| Currency |  | Pending image and source review |
| Issue date |  | Pending image and source review |
| Maturity / call date |  | Pending image and source review |
| Underlying(s) |  | Pending image and source review |
| Coupon / yield |  | Pending image and source review |
| Barrier / protection level |  | Pending image and source review |
| Observation / payment frequency |  | Pending image and source review |
| Redemption terms |  | Pending image and source review |
| Risk notes |  | Pending image and source review |

## Source Evidence

No product terms are inferred from the filename alone. Summary-only records require document recovery or another source before image and OCR evidence can be added.

## Review Log

- Phase 1-2: Placeholder created from the workbook summary.
"""
    path.write_text(content, encoding="utf-8")
    return True


def main() -> None:
    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(WORKBOOK_PATH)
    OUTPUT_DIR.mkdir(exist_ok=True)
    workbook = load_workbook(WORKBOOK_PATH, data_only=True)
    write_summary(workbook)
    isin_sheets = [name for name in workbook.sheetnames if is_isin_sheet(name)]
    summary_isins = {
        str(workbook["ISINs summary"].cell(row, 4).value).strip()
        for row in range(1, workbook["ISINs summary"].max_row + 1)
        if str(workbook["ISINs summary"].cell(row, 4).value or "").strip() in PRODUCT_SUFFIXES
    }
    created = 0
    for isin in sorted(summary_isins):
        old_path = OUTPUT_DIR / f"{isin}.md"
        new_path = OUTPUT_DIR / f"{isin} - {PRODUCT_SUFFIXES[isin]}.md"
        if old_path.exists() and not new_path.exists():
            old_path.rename(new_path)
        created += write_isin_placeholder(workbook, isin)
    print(f"Created {DETAILED_PATH}")
    print(f"Created {SUMMARY_PATH}")
    print(f"Created {created} missing ISIN placeholders in {OUTPUT_DIR}")


if __name__ == "__main__":
    main()
