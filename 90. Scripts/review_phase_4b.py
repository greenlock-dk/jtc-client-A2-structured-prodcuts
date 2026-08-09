from __future__ import annotations

import os
import re
from pathlib import Path

from openpyxl import load_workbook
from generate_phase_1_2 import format_value

ROOT = Path(__file__).resolve().parent.parent
PRODUCT_DIR = ROOT / "01. Structured Products"
OCR_DIR = ROOT / "03. BBG OCR"
REVIEW_DIR = ROOT / "04. Product Review"
WORKBOOK_PATH = ROOT / "Trust ISIN information from Bloomberg.xlsx"

LABELS = {
    "FIGI", "ISIN", "Industry", "ID Number", "Mkt Iss", "Currency", "DBRS", "Fitch",
    "Rank", "Series", "Coupon", "Type", "Cpn Freq", "Iss Price", "Maturity", "Issuer",
    "Name", "Security Information", "Issuer Information", "Identifiers", "Bond Ratings",
}


def link(path: Path, base: Path) -> str:
    return Path(os.path.relpath(path, base)).as_posix().replace(" ", "%20")


def clean(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("|", "\\|").replace("\n", "<br>").strip()


def dossier_for(isin: str) -> Path:
    matches = sorted(PRODUCT_DIR.glob(f"{isin} - *.md"))
    if len(matches) != 1:
        raise FileNotFoundError(f"Expected one dossier for {isin}, found {matches}")
    return matches[0]


def next_value(lines: list[str], label: str) -> str:
    for index, line in enumerate(lines):
        if line.strip().rstrip(":") == label:
            for candidate in lines[index + 1: index + 5]:
                candidate = candidate.strip()
                if candidate and candidate.rstrip(":") not in LABELS and not candidate.startswith(("(", ")")):
                    return candidate
    return ""


def find_candidate(lines: list[str], label: str, pattern: str | None = None) -> str:
    if pattern:
        match = re.search(pattern, "\n".join(lines), re.IGNORECASE)
        return match.group(1).strip() if match else ""
    return next_value(lines, label)


def ocr_fields(isin: str) -> dict[str, str]:
    texts = sorted((OCR_DIR / isin).glob("*.txt"))
    lines = []
    for path in texts:
        lines.extend(line.strip() for line in path.read_text(encoding="utf-8").splitlines() if line.strip())
    combined = "\n".join(lines)
    issuer = find_candidate(lines, "Issuer") or find_candidate(lines, "Name")
    currency = find_candidate(lines, "Currency")
    maturity = find_candidate(lines, "Maturity", r"Maturity\s+([^\n]+)")
    coupons = re.findall(r"\b\d+(?:\.\d+)?%", combined)
    structures = []
    for term in ("Range Accrual", "Dual Range Accrual", "Dual Index", "Ratchet", "Reverse Convertible", "Convertible Bond", "Perpetual", "EURO MTN", "EMTN"):
        if term.lower() in combined.lower() and term not in structures:
            structures.append(term)
    return {
        "Issuer": issuer,
        "Currency": currency,
        "Maturity": maturity,
        "Coupon / yield": ", ".join(dict.fromkeys(coupons)),
        "Product type / structure": ", ".join(structures),
        "ISIN": isin if re.search(re.escape(isin), combined, re.IGNORECASE) else "",
    }


def source_rows(workbook) -> dict[str, dict[str, str]]:
    ws = workbook["ISINs summary"]
    records = {}
    for row in range(7, 16):
        isin = str(ws.cell(row, 4).value or "").strip()
        if isin.startswith(("XS", "CH")):
            records[isin] = {
                "Issuer": "",
                "Currency": "",
                "Maturity": format_value(ws.cell(row, 7)),
                "Coupon / yield": format_value(ws.cell(row, 10)),
                "Product type / structure": clean(ws.cell(row, 9).value),
                "Underlying(s)": clean(ws.cell(row, 14).value),
            }
    for row in range(20, ws.max_row + 1):
        isin = str(ws.cell(row, 4).value or "").strip()
        if isin.startswith(("XS", "CH")):
            records[isin] = {
                "Issuer": clean(ws.cell(row, 7).value),
                "Currency": "",
                "Maturity": format_value(ws.cell(row, 9)),
                "Coupon / yield": format_value(ws.cell(row, 12)),
                "Product type / structure": clean(ws.cell(row, 11).value),
                "Underlying(s)": clean(ws.cell(row, 16).value),
            }
    return records


def write_dossier_section(isin: str, fields: dict[str, str], source: dict[str, str]) -> None:
    dossier = dossier_for(isin)
    ocr_paths = sorted((OCR_DIR / isin).glob("*.txt"))
    evidence = ", ".join(f"[{path.name}]({link(path, dossier.parent)})" for path in ocr_paths)
    rows = []
    for field in ("ISIN", "Issuer", "Currency", "Maturity", "Coupon / yield", "Product type / structure", "Underlying(s)"):
        workbook_value = source.get(field, "")
        candidate = fields.get(field, "")
        if field == "ISIN" and candidate == isin:
            status = "Confirmed: OCR matches worksheet ISIN"
        elif candidate:
            status = "Needs review against source image"
        elif workbook_value:
            status = "Workbook reference; OCR candidate not detected"
        else:
            status = "Not detected"
        rows.append(f"| {field} | {workbook_value} | {candidate} | {status} |")
    section = [
        "## Phase 4B Product Review",
        "",
        "This review layer separates workbook references from OCR candidates. OCR-derived terms are not confirmed until checked visually against the linked source image.",
        "",
        "| Field | Workbook reference | OCR candidate | Review status |",
        "| --- | --- | --- | --- |",
        *rows,
        "",
        f"Raw OCR sources: {evidence}.",
        "",
        "Review status vocabulary: `Confirmed` means the evidence is directly corroborated; `Needs review` requires visual comparison; `Not detected` means OCR did not identify the field.",
    ]
    text = dossier.read_text(encoding="utf-8")
    text = text.replace(
        "Image extraction, OCR transcription, and source-image references will be added in later phases. No product terms are inferred from the worksheet name alone.",
        "Image extraction and raw OCR are complete. Product terms remain subject to the Phase 4B review table and visual confirmation.",
    )
    marker = "## Phase 4B Product Review"
    if marker in text:
        before, after = text.split(marker, 1)
        next_section = after.find("\n## ")
        tail = after[next_section + 1:] if next_section >= 0 else ""
        text = before.rstrip() + "\n\n" + "\n".join(section).rstrip() + "\n"
        if tail:
            text += "\n" + tail.lstrip()
        dossier.write_text(text, encoding="utf-8")
        return
    dossier.write_text(text.rstrip() + "\n\n" + "\n".join(section).rstrip() + "\n", encoding="utf-8")


def main() -> None:
    REVIEW_DIR.mkdir(exist_ok=True)
    workbook = load_workbook(WORKBOOK_PATH, data_only=True)
    source = source_rows(workbook)
    inventory = [
        "# Phase 4B Review Inventory",
        "",
        "> Candidate normalization generated from workbook references and raw OCR. Interpreted terms remain subject to visual review.",
        "",
        "| Record | ISIN | OCR sources | Fields with candidates | Confirmed fields | Review status |",
        "| ---: | --- | ---: | --- | --- | --- |",
    ]
    for record, isin in enumerate(sorted(source), 1):
        candidates = ocr_fields(isin)
        dossier_matches = sorted(PRODUCT_DIR.glob(f"{isin} - *.md"))
        if dossier_matches:
            write_dossier_section(isin, candidates, source[isin])
        candidate_fields = [field for field, value in candidates.items() if value]
        confirmed = ["ISIN"] if candidates.get("ISIN") == isin else []
        dossier_status = "Manual visual review required" if dossier_matches else "No dedicated dossier; worksheet absent"
        inventory.append(
            f"| {record} | `{isin}` | {len(list((OCR_DIR / isin).glob('*.txt')))} | "
            f"{', '.join(candidate_fields) or 'None'} | {', '.join(confirmed) or 'None'} | {dossier_status} |"
        )
    unmapped_lines = [
        "# Unmapped Image Review",
        "",
        "> The following OCR candidates were detected in package media that is not attached to a worksheet drawing. No automatic ISIN assignment has been made.",
        "",
        "| Image | Raw OCR | Candidate identifier | Review status |",
        "| --- | --- | --- | --- |",
    ]
    for image_path in sorted((ROOT / "02. BBG images" / "_unmapped").glob("*.png")):
        ocr_path = OCR_DIR / "_unmapped" / f"{image_path.stem}.txt"
        text = ocr_path.read_text(encoding="utf-8")
        candidates = sorted(set(re.findall(r"(?:XS|CH)[A-Z0-9]{8,14}", text, re.IGNORECASE)))
        candidate_text = ", ".join(candidates) if candidates else "None detected"
        unmapped_lines.append(
            f"| [{image_path.name}]({link(image_path, REVIEW_DIR)}) | [{ocr_path.name}]({link(ocr_path, REVIEW_DIR)}) | "
            f"`{candidate_text}` | Candidate only; visual mapping review required |"
        )
    (REVIEW_DIR / "UNMAPPED REVIEW.md").write_text("\n".join(unmapped_lines) + "\n", encoding="utf-8")
    inventory.extend([
        "",
        "## Review Rules",
        "",
        "- Workbook values are retained as a separate reference layer and are not silently replaced by OCR.",
        "- OCR candidates are not confirmed product terms unless visually checked against the source image.",
        "- The two unmapped images remain outside the ISIN dossiers and require separate identification review.",
    ])
    (REVIEW_DIR / "REVIEW INVENTORY.md").write_text("\n".join(inventory) + "\n", encoding="utf-8")
    print(f"Reviewed {len(source)} ISIN dossiers")
    print(f"Created {REVIEW_DIR / 'REVIEW INVENTORY.md'}")


if __name__ == "__main__":
    main()
