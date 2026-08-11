from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook

from product_frontmatter import read_product, write_product

ROOT = Path(__file__).resolve().parent.parent
WORKBOOK_PATH = ROOT / "Trust ISIN information from Bloomberg.xlsx"
OCR_DIR = ROOT / "03. BBG OCR"
PRODUCT_DIR = ROOT / "01. Structured Products"
CANONICAL_DIR = ROOT / "05. Canonical Data"
SUMMARY_PATH = CANONICAL_DIR / "ISIN_summary.md"
DETAILED_PATH = CANONICAL_DIR / "ISIN_detailed.md"

ISIN_PATTERN = re.compile(r"^(?:XS|CH)\d{10}$")
LABEL_PATTERN = re.compile(
    r"(?i)\b(?:amt\s*issued\s*/\s*outstanding|amount\s+issued\s*/\s*outstanding|"
    r"amount\s+issued|issue\s+size|aggregate\s+(?:nominal|principal|amount)|"
    r"nominal\s+amount|principal\s+amount|outstanding\s+amount)\b"
)
AMOUNT_PATTERN = re.compile(
    r"(?i)\b(USD|EUR|CHF|GBP|AUD|CAD|JPY)\s*\n?\s*([\d,]+(?:\.\d+)?)\s*\(M\)"
)
RECOVERY_MARKER = "## Issue Size / Amount Issued"


def portfolio_isins() -> list[str]:
    workbook = load_workbook(WORKBOOK_PATH, data_only=True, read_only=True)
    worksheet = workbook["ISINs summary"]
    isins = []
    for row in worksheet.iter_rows(values_only=True):
        for value in row:
            if isinstance(value, str) and ISIN_PATTERN.fullmatch(value.strip()) and value.strip() not in isins:
                isins.append(value.strip())
    return isins


def recovery(isin: str) -> dict[str, str]:
    for path in sorted((OCR_DIR / isin).glob("*.txt")):
        lines = path.read_text(encoding="utf-8", errors="replace").splitlines()
        for index, line in enumerate(lines):
            label = LABEL_PATTERN.search(line)
            if not label:
                continue
            window = "\n".join(lines[index + 1 : index + 12])
            amount = AMOUNT_PATTERN.search(window)
            context = " ".join(item.strip() for item in [line, *lines[index + 1 : index + 12]] if item.strip())
            if amount:
                value = f"{amount.group(1).upper()} {amount.group(2)} million"
                return {
                    "value": value,
                    "status": "Candidate: visual confirmation required",
                    "source": f"{path.relative_to(ROOT).as_posix()}; {label.group(0)}; {context}",
                }
            return {
                "value": "Not parsed from OCR",
                "status": "Candidate label: numeric value requires visual confirmation",
                "source": f"{path.relative_to(ROOT).as_posix()}; {label.group(0)}; {context}",
            }
    return {
        "value": "Not available",
        "status": "Unavailable: document recovery required",
        "source": "No issue-size label found in available OCR",
    }


def all_recoveries() -> dict[str, dict[str, str]]:
    values = {isin: recovery(isin) for isin in portfolio_isins()}
    bbva = values["XS3234638248"]
    values["XS3234638248"] = {
        "value": "USD 2.4 million",
        "status": "Confirmed documentary nominal amount",
        "source": "04. Original terms/04. BBVA/XS3234638248_BBVA_Global_Markets_Pricing_Supplement_2026-01-27.pdf",
    }
    values["XS3234638248"]["source"] += "; Trust position size remains USD 2 million"
    return values


def replace_table_columns(path: Path, recoveries: dict[str, dict[str, str]], header: str) -> None:
    lines = path.read_text(encoding="utf-8").splitlines()
    output = []
    table_header_seen = False
    separator_seen = False
    for line in lines:
        if line.startswith("| Record # |") and "Issue/outstanding size" not in line:
            line = line[:-2] + " | Issue/outstanding size | Issue-size status |"
            table_header_seen = True
        elif table_header_seen and line.startswith("| ---") and not separator_seen:
            line = line[:-2] + " | --- | --- |"
            separator_seen = True
        elif table_header_seen and separator_seen and line.startswith("| "):
            match = re.search(r"\b((?:XS|CH)\d{10})\b", line)
            if match and "Issue/outstanding size" not in line:
                value = recoveries[match.group(1)]
                line = line[:-2] + f" | {value['value']} | {value['status']} |"
        output.append(line)
    path.write_text("\n".join(output) + "\n", encoding="utf-8")


def dossier_section(isin: str, value: dict[str, str]) -> str:
    return "\n".join([
        RECOVERY_MARKER,
        "",
        "This field is separate from the Trust's position size. OCR-derived values are candidates until checked against the source image.",
        "",
        "| Field | Value | Status | Source |",
        "| --- | --- | --- | --- |",
        f"| Original issue / amount issued or outstanding | {value['value']} | {value['status']} | `{value['source']}` |",
        "",
    ])


def integrate_dossiers(recoveries: dict[str, dict[str, str]]) -> int:
    updated = 0
    for isin, value in recoveries.items():
        matches = sorted(PRODUCT_DIR.glob(f"{isin} - *.md"))
        if len(matches) != 1:
            continue
        path = matches[0]
        canonical_data, text = read_product(path)
        section = dossier_section(isin, value)
        if RECOVERY_MARKER in text:
            before, after = text.split(RECOVERY_MARKER, 1)
            next_section = after.find("\n## ")
            tail = after[next_section + 1 :] if next_section >= 0 else ""
            text = before.rstrip() + "\n\n" + section
            if tail:
                text += "\n" + tail.lstrip()
        else:
            insertion = text.find("\n## Consistency Review")
            if insertion < 0:
                insertion = text.find("\n## Evidence Sources")
            if insertion < 0:
                text = text.rstrip() + "\n\n" + section
            else:
                text = text[:insertion].rstrip() + "\n\n" + section + text[insertion:]
        if canonical_data:
            canonical_data["issue_size"] = {
                "display": value["value"],
                "status": value["status"],
                "source": value["source"],
            }
            write_product(path, canonical_data, text)
        else:
            path.write_text(text, encoding="utf-8")
        updated += 1
    return updated


def main() -> None:
    recoveries = all_recoveries()
    replace_table_columns(SUMMARY_PATH, recoveries, "summary")
    replace_table_columns(DETAILED_PATH, recoveries, "detailed")
    updated = integrate_dossiers(recoveries)
    print(f"Integrated issue-size fields for {len(recoveries)} instruments")
    print(f"Updated individual dossiers: {updated}")
    print(f"Confirmed documentary amounts: {sum(value['status'].startswith('Confirmed') for value in recoveries.values())}")
    print(f"OCR candidates: {sum(value['status'].startswith('Candidate') for value in recoveries.values())}")
    print(f"Unavailable: {sum(value['status'].startswith('Unavailable') for value in recoveries.values())}")


if __name__ == "__main__":
    main()